import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice


def process_workbook(file_name):
    wb = xl.load_workbook(file_name)
    sheet = wb['Sheet1']

    # lopping over rows
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    sheet['D1'] = 'discounted_price'  # new column title

    # create barchart
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4
                       )

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'f2')
    chart.y_axis.title = 'Discounted price'
    chart.x_axis.title = 'Product ID'
    chart.title = "Sales - Q1"

    series = chart.series[0]
    fill = PatternFillProperties(prst="pct5")
    fill.foreground = ColorChoice(prstClr="red")
    fill.background = ColorChoice(prstClr="blue")
    series.graphicalProperties.pattFill = fill

    wb.save(file_name)


process_workbook('transactions.xlsx')
